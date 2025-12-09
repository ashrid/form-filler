"""
Acknowledgment of Receipt Form GUI
Provides input interface for the acknowledgment form
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import os

from ..pdf.acknowledgment_pdf import AcknowledgmentPDFGenerator

try:
    from openpyxl import load_workbook
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False


class AcknowledgmentFormFrame(ttk.Frame):
    """GUI frame for Acknowledgment of Receipt form"""

    def __init__(self, parent):
        super().__init__(parent)
        self.items = []
        self.item_widgets = []
        self.pdf_generator = AcknowledgmentPDFGenerator()

        self._create_widgets()

    def _create_widgets(self):
        """Create all form widgets"""
        # Main container with scrollbar
        self.canvas = tk.Canvas(self, highlightthickness=0, bg="#f5f5f5")
        self.scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas)

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )

        self.canvas_window = self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        # Make canvas expand to fill width
        self.canvas.bind("<Configure>", self._on_canvas_configure)

        # Pack scrollbar components
        self.scrollbar.pack(side="right", fill="y")
        self.canvas.pack(side="left", fill="both", expand=True)

        # Bind mousewheel only when mouse is over this canvas
        self.canvas.bind("<Enter>", self._bind_mousewheel)
        self.canvas.bind("<Leave>", self._unbind_mousewheel)

        # Form content
        self._create_form_content()

    def _on_canvas_configure(self, event):
        """Update canvas window width when canvas is resized"""
        self.canvas.itemconfig(self.canvas_window, width=event.width)

    def _bind_mousewheel(self, event):
        """Bind mousewheel when mouse enters canvas"""
        # Windows
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)
        # Linux
        self.canvas.bind_all("<Button-4>", self._on_mousewheel_linux)
        self.canvas.bind_all("<Button-5>", self._on_mousewheel_linux)

    def _unbind_mousewheel(self, event):
        """Unbind mousewheel when mouse leaves canvas"""
        self.canvas.unbind_all("<MouseWheel>")
        self.canvas.unbind_all("<Button-4>")
        self.canvas.unbind_all("<Button-5>")

    def _on_mousewheel(self, event):
        """Handle mouse wheel scrolling (Windows)"""
        self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def _on_mousewheel_linux(self, event):
        """Handle mouse wheel scrolling (Linux)"""
        if event.num == 4:
            self.canvas.yview_scroll(-1, "units")
        elif event.num == 5:
            self.canvas.yview_scroll(1, "units")

    def _create_form_content(self):
        """Create the form fields"""
        frame = self.scrollable_frame
        row = 0

        # Title
        title = ttk.Label(frame, text="Acknowledgment of Receipt Form",
                          font=("Helvetica", 16, "bold"))
        title.grid(row=row, column=0, columnspan=6, pady=10)
        row += 1

        # Items section
        items_label = ttk.Label(frame, text="Items:", font=("Helvetica", 12, "bold"))
        items_label.grid(row=row, column=0, sticky="w", padx=5, pady=(10, 5))
        row += 1

        # Items header
        headers = ["Store Code", "Item Description", "Qty", "Purchase Date/LPO", ""]
        for col, header in enumerate(headers):
            lbl = ttk.Label(frame, text=header, font=("Helvetica", 10, "bold"))
            lbl.grid(row=row, column=col, padx=5, pady=2)
        row += 1

        # Items container
        self.items_frame = ttk.Frame(frame)
        self.items_frame.grid(row=row, column=0, columnspan=5, sticky="ew", padx=5)
        self.items_start_row = row
        row += 1

        # Add first item row
        self._add_item_row()

        # Add item button and Import button
        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=row, column=0, columnspan=5, pady=5, sticky="w", padx=5)

        add_btn = ttk.Button(btn_frame, text="+ Add Item", command=self._add_item_row)
        add_btn.pack(side="left", padx=(0, 10))

        import_btn = ttk.Button(btn_frame, text="Import from Excel", command=self._import_from_excel)
        import_btn.pack(side="left")
        row += 1

        # Separator
        ttk.Separator(frame, orient="horizontal").grid(row=row, column=0, columnspan=6, sticky="ew", pady=10)
        row += 1

        # Custodian Details section
        cust_label = ttk.Label(frame, text="Custodian Details:", font=("Helvetica", 12, "bold"))
        cust_label.grid(row=row, column=0, columnspan=2, sticky="w", padx=5, pady=(10, 5))
        row += 1

        # Name
        ttk.Label(frame, text="Name:").grid(row=row, column=0, sticky="e", padx=5, pady=2)
        self.name_entry = ttk.Entry(frame, width=40)
        self.name_entry.grid(row=row, column=1, columnspan=2, sticky="w", padx=5, pady=2)

        ttk.Label(frame, text="Emp. ID:").grid(row=row, column=3, sticky="e", padx=5, pady=2)
        self.emp_id_entry = ttk.Entry(frame, width=15)
        self.emp_id_entry.grid(row=row, column=4, sticky="w", padx=5, pady=2)
        row += 1

        # Department
        ttk.Label(frame, text="College/Department:").grid(row=row, column=0, sticky="e", padx=5, pady=2)
        self.dept_entry = ttk.Entry(frame, width=50)
        self.dept_entry.grid(row=row, column=1, columnspan=3, sticky="w", padx=5, pady=2)
        row += 1

        # Separator
        ttk.Separator(frame, orient="horizontal").grid(row=row, column=0, columnspan=6, sticky="ew", pady=10)
        row += 1

        # Location section
        loc_label = ttk.Label(frame, text="Location:", font=("Helvetica", 12, "bold"))
        loc_label.grid(row=row, column=0, columnspan=2, sticky="w", padx=5, pady=(10, 5))
        row += 1

        # Building
        ttk.Label(frame, text="Building:").grid(row=row, column=0, sticky="e", padx=5, pady=2)
        self.building_var = tk.StringVar()
        buildings = ["SZH", "J1", "J2", "Student Hub", "Hostel", "Others"]
        self.building_combo = ttk.Combobox(frame, textvariable=self.building_var, values=buildings, width=15)
        self.building_combo.grid(row=row, column=1, sticky="w", padx=5, pady=2)

        ttk.Label(frame, text="If Others:").grid(row=row, column=2, sticky="e", padx=5, pady=2)
        self.building_other_entry = ttk.Entry(frame, width=15)
        self.building_other_entry.grid(row=row, column=3, sticky="w", padx=5, pady=2)
        row += 1

        # Floor
        ttk.Label(frame, text="Floor:").grid(row=row, column=0, sticky="e", padx=5, pady=2)
        self.floor_var = tk.StringVar()
        floors = ["Ground", "1st", "2nd", "3rd", "Others"]
        self.floor_combo = ttk.Combobox(frame, textvariable=self.floor_var, values=floors, width=15)
        self.floor_combo.grid(row=row, column=1, sticky="w", padx=5, pady=2)

        ttk.Label(frame, text="If Others:").grid(row=row, column=2, sticky="e", padx=5, pady=2)
        self.floor_other_entry = ttk.Entry(frame, width=15)
        self.floor_other_entry.grid(row=row, column=3, sticky="w", padx=5, pady=2)
        row += 1

        # Section
        ttk.Label(frame, text="Section:").grid(row=row, column=0, sticky="e", padx=5, pady=2)
        self.section_var = tk.StringVar()
        sections = ["Male", "Female"]
        self.section_combo = ttk.Combobox(frame, textvariable=self.section_var, values=sections, width=15)
        self.section_combo.grid(row=row, column=1, sticky="w", padx=5, pady=2)
        row += 1

        # Separator
        ttk.Separator(frame, orient="horizontal").grid(row=row, column=0, columnspan=6, sticky="ew", pady=10)
        row += 1

        # Device Type
        device_label = ttk.Label(frame, text="Device Type:", font=("Helvetica", 12, "bold"))
        device_label.grid(row=row, column=0, columnspan=2, sticky="w", padx=5, pady=(10, 5))
        row += 1

        self.device_var = tk.StringVar(value="Office")
        office_rb = ttk.Radiobutton(frame, text="Office Device", variable=self.device_var, value="Office")
        office_rb.grid(row=row, column=0, columnspan=2, sticky="w", padx=20, pady=2)
        row += 1

        lab_rb = ttk.Radiobutton(frame, text="Lab Device", variable=self.device_var, value="Lab")
        lab_rb.grid(row=row, column=0, columnspan=2, sticky="w", padx=20, pady=2)
        row += 1

        # Separator
        ttk.Separator(frame, orient="horizontal").grid(row=row, column=0, columnspan=6, sticky="ew", pady=10)
        row += 1

        # Generate button
        gen_btn = ttk.Button(frame, text="Generate PDF", command=self._generate_pdf,
                             style="Accent.TButton")
        gen_btn.grid(row=row, column=0, columnspan=6, pady=20)

        # Status label
        row += 1
        self.status_label = ttk.Label(frame, text="")
        self.status_label.grid(row=row, column=0, columnspan=6, pady=5)

    def _add_item_row(self, data: dict = None):
        """Add a new item row, optionally with data"""
        row = len(self.item_widgets)

        store_code = ttk.Entry(self.items_frame, width=15)
        store_code.grid(row=row, column=0, padx=5, pady=2)

        description = ttk.Entry(self.items_frame, width=35)
        description.grid(row=row, column=1, padx=5, pady=2)

        qty = ttk.Entry(self.items_frame, width=8)
        qty.grid(row=row, column=2, padx=5, pady=2)

        purchase_date = ttk.Entry(self.items_frame, width=15)
        purchase_date.grid(row=row, column=3, padx=5, pady=2)

        remove_btn = ttk.Button(self.items_frame, text="X", width=3,
                                command=lambda r=row: self._remove_item_row(r))
        remove_btn.grid(row=row, column=4, padx=5, pady=2)

        # Pre-fill data if provided
        if data:
            store_code.insert(0, data.get("store_code", ""))
            description.insert(0, data.get("description", ""))
            qty.insert(0, data.get("qty", ""))
            purchase_date.insert(0, data.get("purchase_date", ""))

        self.item_widgets.append({
            "store_code": store_code,
            "description": description,
            "qty": qty,
            "purchase_date": purchase_date,
            "remove_btn": remove_btn
        })

    def _import_from_excel(self):
        """Import items from Excel file"""
        if not EXCEL_SUPPORT:
            messagebox.showerror("Error", "Excel support not available.\nPlease install openpyxl: pip install openpyxl")
            return

        filepath = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )

        if not filepath:
            return

        try:
            wb = load_workbook(filepath, data_only=True)
            ws = wb.active

            # Get headers from first row
            headers = [cell.value for cell in ws[1] if cell.value]
            if not headers:
                messagebox.showerror("Error", "No headers found in the Excel file.")
                return

            # Map headers to our fields (case-insensitive)
            header_map = {}
            for i, h in enumerate(headers):
                h_lower = str(h).lower().strip()
                if "store" in h_lower or "code" in h_lower:
                    header_map["store_code"] = i
                elif "description" in h_lower or "item" in h_lower:
                    header_map["description"] = i
                elif "qty" in h_lower or "quantity" in h_lower:
                    header_map["qty"] = i
                elif "date" in h_lower or "lpo" in h_lower or "purchase" in h_lower:
                    header_map["purchase_date"] = i

            # Clear existing items (except first row)
            for widgets in self.item_widgets[1:]:
                for widget in widgets.values():
                    widget.destroy()
            self.item_widgets = self.item_widgets[:1]

            # Clear first row
            for key in ["store_code", "description", "qty", "purchase_date"]:
                self.item_widgets[0][key].delete(0, tk.END)

            # Read data rows
            items_added = 0
            for row_num, row in enumerate(ws.iter_rows(min_row=2), 2):
                row_values = [cell.value for cell in row]

                # Skip empty rows
                if not any(row_values):
                    continue

                item_data = {
                    "store_code": str(row_values[header_map.get("store_code", 0)] or ""),
                    "description": str(row_values[header_map.get("description", 1)] or ""),
                    "qty": str(row_values[header_map.get("qty", 2)] or ""),
                    "purchase_date": str(row_values[header_map.get("purchase_date", 3)] or ""),
                }

                if items_added == 0:
                    # Fill first row
                    self.item_widgets[0]["store_code"].insert(0, item_data["store_code"])
                    self.item_widgets[0]["description"].insert(0, item_data["description"])
                    self.item_widgets[0]["qty"].insert(0, item_data["qty"])
                    self.item_widgets[0]["purchase_date"].insert(0, item_data["purchase_date"])
                else:
                    # Add new row
                    self._add_item_row(item_data)

                items_added += 1

            messagebox.showinfo("Success", f"Imported {items_added} items from Excel.")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to import Excel file:\n{str(e)}")

    def _remove_item_row(self, row_index):
        """Remove an item row"""
        if len(self.item_widgets) <= 1:
            messagebox.showwarning("Warning", "At least one item row is required.")
            return

        widgets = self.item_widgets[row_index]
        for widget in widgets.values():
            widget.destroy()

        self.item_widgets.pop(row_index)

        # Reorganize remaining rows
        for i, widgets in enumerate(self.item_widgets):
            widgets["store_code"].grid(row=i, column=0)
            widgets["description"].grid(row=i, column=1)
            widgets["qty"].grid(row=i, column=2)
            widgets["purchase_date"].grid(row=i, column=3)
            widgets["remove_btn"].grid(row=i, column=4)
            widgets["remove_btn"].configure(command=lambda r=i: self._remove_item_row(r))

    def _get_items(self) -> list:
        """Get all items from the form"""
        items = []
        for widgets in self.item_widgets:
            store_code = widgets["store_code"].get().strip()
            description = widgets["description"].get().strip()
            qty = widgets["qty"].get().strip()
            purchase_date = widgets["purchase_date"].get().strip()

            if store_code or description or qty or purchase_date:
                items.append({
                    "store_code": store_code,
                    "description": description,
                    "qty": qty,
                    "purchase_date": purchase_date
                })
        return items

    def _generate_pdf(self):
        """Generate the PDF form"""
        items = self._get_items()

        if not items:
            messagebox.showwarning("Warning", "Please add at least one item.")
            return

        form_data = {
            "items": items,
            "custodian_name": self.name_entry.get().strip(),
            "emp_id": self.emp_id_entry.get().strip(),
            "department": self.dept_entry.get().strip(),
            "building": self.building_var.get(),
            "building_other": self.building_other_entry.get().strip(),
            "floor": self.floor_var.get(),
            "floor_other": self.floor_other_entry.get().strip(),
            "section": self.section_var.get(),
            "device_type": self.device_var.get()
        }

        try:
            filepath = self.pdf_generator.generate(form_data)
            self.status_label.config(text=f"PDF generated: {filepath}", foreground="green")
            messagebox.showinfo("Success", f"PDF generated successfully!\n\n{filepath}")

            # Open the PDF
            if os.name == 'nt':
                os.startfile(filepath)
            elif os.name == 'posix':
                os.system(f'xdg-open "{filepath}"')
        except Exception as e:
            self.status_label.config(text=f"Error: {str(e)}", foreground="red")
            messagebox.showerror("Error", f"Failed to generate PDF:\n{str(e)}")

    def clear_form(self):
        """Clear all form fields"""
        # Clear items
        for widgets in self.item_widgets[1:]:
            for widget in widgets.values():
                widget.destroy()
        self.item_widgets = self.item_widgets[:1]

        # Clear first item
        for key in ["store_code", "description", "qty", "purchase_date"]:
            self.item_widgets[0][key].delete(0, tk.END)

        # Clear other fields
        self.name_entry.delete(0, tk.END)
        self.emp_id_entry.delete(0, tk.END)
        self.dept_entry.delete(0, tk.END)
        self.building_var.set("")
        self.building_other_entry.delete(0, tk.END)
        self.floor_var.set("")
        self.floor_other_entry.delete(0, tk.END)
        self.section_var.set("")
        self.device_var.set("Office")
        self.status_label.config(text="")
