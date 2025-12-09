"""
Asset Transfer Form (ATF) GUI
Provides input interface for the asset transfer form
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import os

from ..pdf.transfer_pdf import TransferPDFGenerator

try:
    from openpyxl import load_workbook
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False


class TransferFormFrame(ttk.Frame):
    """GUI frame for Asset Transfer Form"""

    def __init__(self, parent):
        super().__init__(parent)
        self.assets = []
        self.asset_widgets = []
        self.pdf_generator = TransferPDFGenerator()

        self._create_widgets()

    def _create_widgets(self):
        """Create all form widgets"""
        # Main container with scrollbar
        self.canvas = tk.Canvas(self, highlightthickness=0, bg="#f5f5f5")
        self.scrollbar = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas)

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )

        self.canvas_window = self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        # Make canvas expand to fill width
        self.canvas.bind("<Configure>", self._on_canvas_configure)

        # Pack scrollbar components
        self.scrollbar.pack(side="right", fill="y")
        self.canvas.pack(side="left", fill="both", expand=True)

        # Bind mousewheel only when mouse is over this canvas
        self.canvas.bind("<Enter>", self._bind_mousewheel)
        self.canvas.bind("<Leave>", self._unbind_mousewheel)

        # Form content
        self._create_form_content()

    def _on_canvas_configure(self, event):
        """Update canvas window width when canvas is resized"""
        self.canvas.itemconfig(self.canvas_window, width=event.width)

    def _bind_mousewheel(self, event):
        """Bind mousewheel when mouse enters canvas"""
        # Windows
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)
        # Linux
        self.canvas.bind_all("<Button-4>", self._on_mousewheel_linux)
        self.canvas.bind_all("<Button-5>", self._on_mousewheel_linux)

    def _unbind_mousewheel(self, event):
        """Unbind mousewheel when mouse leaves canvas"""
        self.canvas.unbind_all("<MouseWheel>")
        self.canvas.unbind_all("<Button-4>")
        self.canvas.unbind_all("<Button-5>")

    def _on_mousewheel(self, event):
        """Handle mouse wheel scrolling (Windows)"""
        self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def _on_mousewheel_linux(self, event):
        """Handle mouse wheel scrolling (Linux)"""
        if event.num == 4:
            self.canvas.yview_scroll(-1, "units")
        elif event.num == 5:
            self.canvas.yview_scroll(1, "units")

    def _create_form_content(self):
        """Create the form fields"""
        frame = self.scrollable_frame
        row = 0

        # Title
        title = ttk.Label(frame, text="Asset Transfer Form (ATF)",
                          font=("Helvetica", 16, "bold"))
        title.grid(row=row, column=0, columnspan=6, pady=10)
        row += 1

        # Transferred From section
        from_label = ttk.Label(frame, text="Transferred From:", font=("Helvetica", 12, "bold"))
        from_label.grid(row=row, column=0, columnspan=2, sticky="w", padx=5, pady=(10, 5))
        row += 1

        # From - Custodian Name
        ttk.Label(frame, text="Custodian Name:").grid(row=row, column=0, sticky="e", padx=5, pady=2)
        self.from_name_entry = ttk.Entry(frame, width=40)
        self.from_name_entry.grid(row=row, column=1, columnspan=2, sticky="w", padx=5, pady=2)
        row += 1

        # From - Department
        ttk.Label(frame, text="Department:").grid(row=row, column=0, sticky="e", padx=5, pady=2)
        self.from_dept_entry = ttk.Entry(frame, width=40)
        self.from_dept_entry.grid(row=row, column=1, columnspan=2, sticky="w", padx=5, pady=2)

        ttk.Label(frame, text="Emp. ID:").grid(row=row, column=3, sticky="e", padx=5, pady=2)
        self.from_emp_id_entry = ttk.Entry(frame, width=15)
        self.from_emp_id_entry.grid(row=row, column=4, sticky="w", padx=5, pady=2)
        row += 1

        # Separator
        ttk.Separator(frame, orient="horizontal").grid(row=row, column=0, columnspan=6, sticky="ew", pady=10)
        row += 1

        # Assets section
        assets_label = ttk.Label(frame, text="Assets:", font=("Helvetica", 12, "bold"))
        assets_label.grid(row=row, column=0, columnspan=2, sticky="w", padx=5, pady=(10, 5))
        row += 1

        # Assets header
        headers = ["Store Code", "Asset Name", "Description", "Old Asset No.", ""]
        for col, header in enumerate(headers):
            lbl = ttk.Label(frame, text=header, font=("Helvetica", 10, "bold"))
            lbl.grid(row=row, column=col, padx=5, pady=2)
        row += 1

        # Assets container
        self.assets_frame = ttk.Frame(frame)
        self.assets_frame.grid(row=row, column=0, columnspan=5, sticky="ew", padx=5)
        row += 1

        # Add first asset row
        self._add_asset_row()

        # Add asset button and Import button
        btn_frame = ttk.Frame(frame)
        btn_frame.grid(row=row, column=0, columnspan=5, pady=5, sticky="w", padx=5)

        add_btn = ttk.Button(btn_frame, text="+ Add Asset", command=self._add_asset_row)
        add_btn.pack(side="left", padx=(0, 10))

        import_btn = ttk.Button(btn_frame, text="Import from Excel", command=self._import_from_excel)
        import_btn.pack(side="left")
        row += 1

        # Separator
        ttk.Separator(frame, orient="horizontal").grid(row=row, column=0, columnspan=6, sticky="ew", pady=10)
        row += 1

        # Transferred To section
        to_label = ttk.Label(frame, text="Transferred To:", font=("Helvetica", 12, "bold"))
        to_label.grid(row=row, column=0, columnspan=2, sticky="w", padx=5, pady=(10, 5))
        row += 1

        # To - Custodian Name
        ttk.Label(frame, text="Custodian Name:").grid(row=row, column=0, sticky="e", padx=5, pady=2)
        self.to_name_entry = ttk.Entry(frame, width=40)
        self.to_name_entry.grid(row=row, column=1, columnspan=2, sticky="w", padx=5, pady=2)
        row += 1

        # To - Department and Emp ID
        ttk.Label(frame, text="Department:").grid(row=row, column=0, sticky="e", padx=5, pady=2)
        self.to_dept_entry = ttk.Entry(frame, width=40)
        self.to_dept_entry.grid(row=row, column=1, columnspan=2, sticky="w", padx=5, pady=2)

        ttk.Label(frame, text="Emp. ID:").grid(row=row, column=3, sticky="e", padx=5, pady=2)
        self.to_emp_id_entry = ttk.Entry(frame, width=15)
        self.to_emp_id_entry.grid(row=row, column=4, sticky="w", padx=5, pady=2)
        row += 1

        # Separator
        ttk.Separator(frame, orient="horizontal").grid(row=row, column=0, columnspan=6, sticky="ew", pady=10)
        row += 1

        # Generate button
        gen_btn = ttk.Button(frame, text="Generate PDF", command=self._generate_pdf,
                             style="Accent.TButton")
        gen_btn.grid(row=row, column=0, columnspan=6, pady=20)

        # Status label
        row += 1
        self.status_label = ttk.Label(frame, text="")
        self.status_label.grid(row=row, column=0, columnspan=6, pady=5)

    def _add_asset_row(self, data: dict = None):
        """Add a new asset row, optionally with data"""
        row = len(self.asset_widgets)

        store_code = ttk.Entry(self.assets_frame, width=12)
        store_code.grid(row=row, column=0, padx=5, pady=2)

        asset_name = ttk.Entry(self.assets_frame, width=18)
        asset_name.grid(row=row, column=1, padx=5, pady=2)

        description = ttk.Entry(self.assets_frame, width=25)
        description.grid(row=row, column=2, padx=5, pady=2)

        old_asset_no = ttk.Entry(self.assets_frame, width=15)
        old_asset_no.grid(row=row, column=3, padx=5, pady=2)

        remove_btn = ttk.Button(self.assets_frame, text="X", width=3,
                                command=lambda r=row: self._remove_asset_row(r))
        remove_btn.grid(row=row, column=4, padx=5, pady=2)

        # Pre-fill data if provided
        if data:
            store_code.insert(0, data.get("store_code", ""))
            asset_name.insert(0, data.get("asset_name", ""))
            description.insert(0, data.get("description", ""))
            old_asset_no.insert(0, data.get("old_asset_no", ""))

        self.asset_widgets.append({
            "store_code": store_code,
            "asset_name": asset_name,
            "description": description,
            "old_asset_no": old_asset_no,
            "remove_btn": remove_btn
        })

    def _import_from_excel(self):
        """Import assets from Excel file"""
        if not EXCEL_SUPPORT:
            messagebox.showerror("Error", "Excel support not available.\nPlease install openpyxl: pip install openpyxl")
            return

        filepath = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )

        if not filepath:
            return

        try:
            wb = load_workbook(filepath, data_only=True)
            ws = wb.active

            # Get headers from first row
            headers = [cell.value for cell in ws[1] if cell.value]
            if not headers:
                messagebox.showerror("Error", "No headers found in the Excel file.")
                return

            # Map headers to our fields (case-insensitive)
            header_map = {}
            for i, h in enumerate(headers):
                h_lower = str(h).lower().strip()
                if "store" in h_lower or "code" in h_lower:
                    header_map["store_code"] = i
                elif "asset" in h_lower and "name" in h_lower:
                    header_map["asset_name"] = i
                elif "description" in h_lower or "desc" in h_lower:
                    header_map["description"] = i
                elif "old" in h_lower or "asset no" in h_lower:
                    header_map["old_asset_no"] = i

            # Clear existing assets (except first row)
            for widgets in self.asset_widgets[1:]:
                for widget in widgets.values():
                    widget.destroy()
            self.asset_widgets = self.asset_widgets[:1]

            # Clear first row
            for key in ["store_code", "asset_name", "description", "old_asset_no"]:
                self.asset_widgets[0][key].delete(0, tk.END)

            # Read data rows
            assets_added = 0
            for row_num, row in enumerate(ws.iter_rows(min_row=2), 2):
                row_values = [cell.value for cell in row]

                # Skip empty rows
                if not any(row_values):
                    continue

                asset_data = {
                    "store_code": str(row_values[header_map.get("store_code", 0)] or ""),
                    "asset_name": str(row_values[header_map.get("asset_name", 1)] or ""),
                    "description": str(row_values[header_map.get("description", 2)] or ""),
                    "old_asset_no": str(row_values[header_map.get("old_asset_no", 3)] or ""),
                }

                if assets_added == 0:
                    # Fill first row
                    self.asset_widgets[0]["store_code"].insert(0, asset_data["store_code"])
                    self.asset_widgets[0]["asset_name"].insert(0, asset_data["asset_name"])
                    self.asset_widgets[0]["description"].insert(0, asset_data["description"])
                    self.asset_widgets[0]["old_asset_no"].insert(0, asset_data["old_asset_no"])
                else:
                    # Add new row
                    self._add_asset_row(asset_data)

                assets_added += 1

            messagebox.showinfo("Success", f"Imported {assets_added} assets from Excel.")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to import Excel file:\n{str(e)}")

    def _remove_asset_row(self, row_index):
        """Remove an asset row"""
        if len(self.asset_widgets) <= 1:
            messagebox.showwarning("Warning", "At least one asset row is required.")
            return

        widgets = self.asset_widgets[row_index]
        for widget in widgets.values():
            widget.destroy()

        self.asset_widgets.pop(row_index)

        # Reorganize remaining rows
        for i, widgets in enumerate(self.asset_widgets):
            widgets["store_code"].grid(row=i, column=0)
            widgets["asset_name"].grid(row=i, column=1)
            widgets["description"].grid(row=i, column=2)
            widgets["old_asset_no"].grid(row=i, column=3)
            widgets["remove_btn"].grid(row=i, column=4)
            widgets["remove_btn"].configure(command=lambda r=i: self._remove_asset_row(r))

    def _get_assets(self) -> list:
        """Get all assets from the form"""
        assets = []
        for widgets in self.asset_widgets:
            store_code = widgets["store_code"].get().strip()
            asset_name = widgets["asset_name"].get().strip()
            description = widgets["description"].get().strip()
            old_asset_no = widgets["old_asset_no"].get().strip()

            if store_code or asset_name or description or old_asset_no:
                assets.append({
                    "store_code": store_code,
                    "asset_name": asset_name,
                    "description": description,
                    "old_asset_no": old_asset_no
                })
        return assets

    def _generate_pdf(self):
        """Generate the PDF form"""
        assets = self._get_assets()

        if not assets:
            messagebox.showwarning("Warning", "Please add at least one asset.")
            return

        form_data = {
            "from_name": self.from_name_entry.get().strip(),
            "from_department": self.from_dept_entry.get().strip(),
            "from_emp_id": self.from_emp_id_entry.get().strip(),
            "to_name": self.to_name_entry.get().strip(),
            "to_department": self.to_dept_entry.get().strip(),
            "to_emp_id": self.to_emp_id_entry.get().strip(),
            "assets": assets
        }

        try:
            filepath = self.pdf_generator.generate(form_data)
            self.status_label.config(text=f"PDF generated: {filepath}", foreground="green")
            messagebox.showinfo("Success", f"PDF generated successfully!\n\n{filepath}")

            # Open the PDF
            if os.name == 'nt':
                os.startfile(filepath)
            elif os.name == 'posix':
                os.system(f'xdg-open "{filepath}"')
        except Exception as e:
            self.status_label.config(text=f"Error: {str(e)}", foreground="red")
            messagebox.showerror("Error", f"Failed to generate PDF:\n{str(e)}")

    def clear_form(self):
        """Clear all form fields"""
        # Clear assets
        for widgets in self.asset_widgets[1:]:
            for widget in widgets.values():
                widget.destroy()
        self.asset_widgets = self.asset_widgets[:1]

        # Clear first asset
        for key in ["store_code", "asset_name", "description", "old_asset_no"]:
            self.asset_widgets[0][key].delete(0, tk.END)

        # Clear other fields
        self.from_name_entry.delete(0, tk.END)
        self.from_dept_entry.delete(0, tk.END)
        self.from_emp_id_entry.delete(0, tk.END)
        self.to_name_entry.delete(0, tk.END)
        self.to_dept_entry.delete(0, tk.END)
        self.to_emp_id_entry.delete(0, tk.END)
        self.status_label.config(text="")
